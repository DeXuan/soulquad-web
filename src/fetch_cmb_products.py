#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""招行APP理财产品收益率对比表生成器"""

import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 产品代码前缀 -> 机构名称
PREFIX_MAP = {
    "GD": "光大理财", "XY": "兴银理财", "HX": "华夏理财",
    "ZY": "中银理财", "ZX": "中信理财", "PX": "平安理财",
    "NY": "农银理财", "JY": "交银理财", "YC": "渝农商理财",
    "PY": "浦银理财", "MS": "民生理财",
}

# 所有产品代码
PRODUCT_CODES = [
    "GD040325", "XY040226", "HX040211", "ZY040305", "ZX040209",
    "133037A", "HX040214", "PX040218", "NY040206", "JY040231",
    "17385D", "GD40203", "11342P", "17356D", "JY040225",
    "ZY040214", "ZX040228", "HX040217", "PY040217", "133038A",
    "YC040310", "NY040209", "120057A", "PY040207", "ZX040310",
    "YC040309", "GD040315", "134016A", "GD040216", "ZY040209",
    "120082A", "132031D", "YC040211", "MS040207", "GD040230",
    "XY040224",
]

def get_prefix(code):
    for prefix in PREFIX_MAP:
        if code.startswith(prefix):
            return prefix
    return None

def get_institution(code):
    prefix = get_prefix(code)
    if prefix:
        return PREFIX_MAP[prefix]
    if code[0].isdigit():
        return "银行代销基金"
    return "未知"

# 按机构分组
groups = {}
for code in PRODUCT_CODES:
    inst = get_institution(code)
    if inst not in groups:
        groups[inst] = []
    groups[inst].append(code)

# 机构排序
INST_ORDER = [
    "光大理财", "兴银理财", "华夏理财", "中银理财", "中信理财",
    "平安理财", "农银理财", "交银理财", "渝农商理财", "浦银理财",
    "民生理财", "银行代销基金",
]

# ============================================================
# 生成Excel
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "招行理财产品收益对比"

# 样式
hdr_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
hdr_fill_blue = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
hdr_fill_green = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
data_font = Font(name='Microsoft YaHei', size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 机构颜色
COLORS = {
    "光大理财": "FFF2CC", "兴银理财": "D9E2F3", "华夏理财": "E2EFDA",
    "中银理财": "FCE4D6", "中信理财": "D6DCE4", "平安理财": "F8CBAD",
    "农银理财": "D5A6BD", "交银理财": "B4C7E7", "渝农商理财": "C6EFCE",
    "浦银理财": "FFE699", "民生理财": "D9D2E9", "银行代销基金": "DDEBF7",
}

# 标题
ws.merge_cells('A1:H1')
title = ws['A1']
title.value = '招商银行APP理财产品收益率对比表'
title.font = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496')
title.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

# 日期行
ws.merge_cells('A2:H2')
date_cell = ws['A2']
date_cell.value = '数据日期: 请在招行APP中查看最新收益率后填入下方 (收益率单位: %)'
date_cell.font = Font(name='Microsoft YaHei', size=10, color='666666', italic=True)
date_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# 表头
HEADERS = [
    ('A', '产品代码', 14), ('B', '发行机构', 14), ('C', '产品名称(可选)', 28),
    ('D', '近1年(%)', 12), ('E', '近6月(%)', 12), ('F', '近3月(%)', 12),
    ('G', '近1月(%)', 12), ('H', '近1周(%)', 12),
]
for col_letter, hdr_text, width in HEADERS:
    cell = ws[f'{col_letter}3']
    cell.value = hdr_text
    cell.font = hdr_font
    cell.alignment = center
    ws.column_dimensions[col_letter].width = width
    cell.fill = hdr_fill_green if col_letter in 'DEFGH' else hdr_fill_blue
ws.row_dimensions[3].height = 30

# 数据行
row = 4
for inst in INST_ORDER:
    if inst not in groups:
        continue
    codes = groups[inst]
    color = COLORS.get(inst, "F2F2F2")
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    # 机构分组标题
    ws.merge_cells(f'A{row}:H{row}')
    grp = ws[f'A{row}']
    grp.value = f"  {inst} ({len(codes)}款)"
    grp.font = Font(name='Microsoft YaHei', bold=True, size=11, color='2F5496')
    grp.alignment = left
    grp.fill = fill
    ws.row_dimensions[row].height = 24
    row += 1

    # 每个产品一行
    for code in codes:
        cells = [('A', code), ('B', inst), ('C', ''), ('D', ''), ('E', ''), ('F', ''), ('G', ''), ('H', '')]
        for col_letter, val in cells:
            cell = ws[f'{col_letter}{row}']
            cell.value = val
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center if col_letter != 'C' else left
        ws.row_dimensions[row].height = 22
        row += 1

# 说明区
row += 1
ws.merge_cells(f'A{row}:H{row}')
note = ws[f'A{row}']
note.value = (
    '【使用说明】\n'
    '1. 打开招商银行APP → 理财 → 搜索产品代码 → 查看各期限收益率\n'
    '2. 将收益率数字直接填入对应单元格(填数字即可，如3.25)\n'
    '3. 产品代码前缀含义: GD=光大理财, XY=兴银理财, HX=华夏理财, ZY=中银理财, ZX=中信理财, '
    'PX=平安理财, NY=农银理财, JY=交银理财, YC=渝农商理财, PY=浦银理财, MS=民生理财\n'
    '4. 数字开头的代码(如133037A)为银行代销基金/理财产品，同样在招行APP中搜索查看'
)
note.font = Font(name='Microsoft YaHei', size=9, color='666666')
note.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.row_dimensions[row].height = 80

output_path = r'C:\Users\gdx\Desktop\招行理财产品收益对比表.xlsx'
wb.save(output_path)
print(f'Excel saved to: {output_path}')
print(f'共 {len(PRODUCT_CODES)} 个产品，分 {len(groups)} 个机构组')
print('请在招行APP中查看收益率后手动填入Excel')
