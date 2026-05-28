#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
沪深300指数基金分析 - 保存到桌面
数据来源: 天天基金网 (2026-05-27)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ============ Sheet 1: 被动指数基金 ============
ws = wb.active
ws.title = "被动指数基金"

# 标题行
title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws.merge_cells('A1:H1')
ws['A1'] = '沪深300被动指数基金对比分析 (数据截至 2026-05-27)'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# 表头
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
headers = ['基金代码', '基金名称', '管理费率(%/年)', '托管费率(%/年)', '近1年(%)', '近6月(%)', '近3月(%)', '近1周(%)']

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 35

# 数据
passive_data = [
    ['110020', '易方达沪深300ETF联接A', 0.15, 0.05, 29.96, 9.74, 4.73, 1.87],
    ['000051', '华夏沪深300ETF联接A', 0.15, 0.05, 29.87, 9.71, 4.66, 1.88],
    ['160706', '嘉实沪深300ETF联接A', 0.15, 0.05, 29.57, 9.54, 4.65, 1.88],
    ['510310', '沪深300ETF易方达', 0.15, 0.05, 31.70, 10.22, 4.92, 1.98],
    ['510300', '沪深300ETF华泰柏瑞', 0.15, 0.05, 31.47, 10.10, 4.86, 1.99],
    ['510330', '沪深300ETF华夏', 0.15, 0.05, 31.52, 10.14, 4.90, 1.99],
    ['270010', '广发沪深300ETF联接A', 0.50, 0.10, 28.66, 9.33, 4.56, 1.88],
    ['000656', '前海开源沪深300指数A', 0.50, 0.10, 30.55, 10.91, 5.26, 1.96],
    ['005102', '工银沪深300ETF联接A', 0.15, 0.05, 29.21, 9.24, 4.77, 1.87],
]

# 颜色定义
light_blue = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
gold_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

for row_idx, fund in enumerate(passive_data, 3):
    fill = light_blue if row_idx % 2 == 1 else white_fill
    for col_idx, value in enumerate(fund, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='微软雅黑', size=10)
        # 收益率列高亮
        if col_idx >= 5 and isinstance(value, (int, float)):
            if value > 30:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='FF0000')
            elif value > 10:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='FF6600')
            elif value > 0:
                cell.font = Font(name='微软雅黑', size=10, color='008000')

# 设置列宽
for i, width in enumerate([12, 28, 15, 15, 12, 12, 12, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ============ Sheet 2: 指数增强基金 ============
ws2 = wb.create_sheet("指数增强基金")

ws2.merge_cells('A1:I1')
ws2['A1'] = '沪深300指数增强基金对比分析 (数据截至 2026-05-27)'
ws2['A1'].font = title_font
ws2['A1'].fill = title_fill
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 40

headers2 = ['基金代码', '基金名称', '管理费率(%/年)', '托管费率(%/年)', '近1年(%)', '近6月(%)', '近3月(%)', '近1周(%)', '超额收益(%)']
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws2.row_dimensions[2].height = 35

enhanced_data = [
    ['005113', '平安沪深300指数量化A', 1.00, 0.10, 44.70, 16.74, 5.56, 1.71, 14.74],
    ['000368', '汇添富沪深300安中指数A', 0.50, 0.10, 39.71, 12.94, 6.26, 1.55, 9.75],
    ['003884', '汇安沪深300指数增强A', 1.00, 0.10, 37.25, 14.12, 7.09, 1.99, 7.29],
    ['003548', '宏利沪深300指数增强C', 0.65, 0.12, 31.26, 15.73, 6.35, 3.11, 1.30],
    ['001015', '华夏沪深300指数增强A', 1.00, 0.20, 30.89, 13.25, 7.96, 2.53, 0.93],
    ['002310', '创金合信沪深300指数增强A', 0.80, 0.10, 24.85, 6.98, 3.27, 0.55, -5.11],
    ['003885', '汇安沪深300指数增强C', 1.00, 0.10, 36.71, 13.89, 6.98, 1.99, 6.75],
    ['005114', '平安沪深300指数量化C', 1.00, 0.10, 43.98, 16.44, 5.43, 1.71, 14.02],
]

light_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

for row_idx, fund in enumerate(enhanced_data, 3):
    fill = light_green if row_idx % 2 == 1 else white_fill
    for col_idx, value in enumerate(fund, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='微软雅黑', size=10)
        # 超额收益高亮
        if col_idx == 9 and isinstance(value, (int, float)):
            if value > 10:
                cell.font = Font(name='微软雅黑', size=11, bold=True, color='FF0000')
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            elif value > 5:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='FF6600')
            elif value > 0:
                cell.font = Font(name='微软雅黑', size=10, color='008000')
            else:
                cell.font = Font(name='微软雅黑', size=10, color='FF0000')

for i, width in enumerate([12, 28, 15, 15, 12, 12, 12, 12, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ============ Sheet 3: 投资建议 ============
ws3 = wb.create_sheet("投资建议")

ws3.merge_cells('A1:F1')
ws3['A1'] = '沪深300基金投资建议与买入时机'
ws3['A1'].font = title_font
ws3['A1'].fill = title_fill
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

# 当前市场分析
ws3.merge_cells('A3:F3')
ws3['A3'] = '一、当前市场环境分析'
ws3['A3'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')

market_data = [
    ['指标', '数值', '判断', '说明'],
    ['当前点位', '4947.85', '中位偏低', '距离历史高点仍有空间'],
    ['近3月涨幅', '+12%', '上升趋势', '从3月低点4418反弹'],
    ['市盈率(PE)', '~12倍', '估值合理', '处于历史中位数偏低'],
    ['近1周涨幅', '+1.9%', '短期向好', '突破前期震荡区间'],
]

for row_idx, data in enumerate(market_data, 4):
    for col_idx, value in enumerate(data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 4:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = light_blue if row_idx % 2 == 0 else white_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 推荐组合
ws3.merge_cells('A10:F10')
ws3['A10'] = '二、推荐基金组合'
ws3['A10'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')

portfolio = [
    ['风险偏好', '推荐组合', '配置比例', '预期3个月收益', '理由', '适合人群'],
    ['稳健型', '110020 + 005113', '60% + 40%', '8-12%', '低费率打底+量化增强', '保守投资者'],
    ['进取型', '003548 + 001015', '50% + 50%', '10-15%', '近期动量强，增强效果显著', '有经验投资者'],
    ['定投型', '000368', '100%', '8-12%', '费率低，收益稳定', '定投用户'],
    ['高收益型', '005113', '100%', '12-18%', '量化增强效果最佳', '激进投资者'],
]

for row_idx, data in enumerate(portfolio, 11):
    for col_idx, value in enumerate(data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 11:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = light_green if row_idx % 2 == 1 else white_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 操作建议
ws3.merge_cells('A17:F17')
ws3['A17'] = '三、操作建议'
ws3['A17'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')

tips = [
    '1. 买入时机：沪深300刚突破4950，趋势向上，建议分批建仓',
    '2. 定投频率：每周一次，避免追高，平滑成本',
    '3. 止盈点位：若3个月内达到5200-5300点（+5-7%），可部分止盈',
    '4. 止损策略：若跌破4800点，考虑加仓而非止损（强支撑位）',
    '5. 持有周期：建议至少持有3-6个月，享受经济复苏红利',
    '6. 风险控制：单只基金仓位不超过总资产的40%',
]

for row_idx, tip in enumerate(tips, 18):
    ws3.merge_cells(f'A{row_idx}:F{row_idx}')
    cell = ws3.cell(row=row_idx, column=1, value=tip)
    cell.font = Font(name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

for i, width in enumerate([15, 22, 15, 18, 30, 15], 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

# 保存到桌面
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
output_path = os.path.join(desktop_path, '沪深300指数基金分析_20260527.xlsx')
wb.save(output_path)
print(f"Excel文件已保存到桌面: {output_path}")
