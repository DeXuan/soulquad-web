#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
苹果概念基金分析 - 支付宝可购买基金
数据来源: 天天基金网 (2026-05-26)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ Sheet 1: 基金对比表 ============
ws = wb.active
ws.title = "苹果概念基金对比"

# 标题行
title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ws.merge_cells('A1:L1')
ws['A1'] = '支付宝苹果概念基金对比分析 (数据截至 2026-05-26)'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# 表头
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
headers = [
    '基金代码', '基金名称', '基金类型', '跟踪标的',
    '管理费率(%/年)', '托管费率(%/年)', '申购费率(支付宝)',
    '近1年(%)', '近6月(%)', '近3月(%)', '近1月(%)', '近1周(%)'
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 35

# 数据
funds_data = [
    # 被动指数型基金 (跟踪纳斯达克100/标普500)
    ['270042', '广发纳斯达克100ETF联接A', '指数型-海外股票', '纳斯达克100', 0.80, 0.20, '0.13%', 31.04, 17.12, 17.22, 8.92, 1.47],
    ['006479', '广发纳斯达克100ETF联接C', '指数型-海外股票', '纳斯达克100', 0.80, 0.20, '0费率', 30.78, 17.01, 17.16, 8.91, 1.47],
    ['040046', '华安纳斯达克100ETF联接A', '指数型-海外股票', '纳斯达克100', 0.60, 0.20, '0.12%', 32.68, 13.52, 14.65, 7.46, 1.52],
    ['015299', '华夏纳斯达克100ETF联接A', '指数型-海外股票', '纳斯达克100', 0.60, 0.20, '0.12%', 31.33, 12.60, 14.57, 6.94, 1.36],
    ['000834', '大成纳斯达克100ETF联接A', '指数型-海外股票', '纳斯达克100', 0.80, 0.20, '0.12%', 31.01, 16.55, 16.54, 8.38, 1.39],
    ['019547', '招商纳斯达克100ETF联接A', '指数型-海外股票', '纳斯达克100', 0.50, 0.15, '0.12%', 30.83, 12.75, 13.85, 7.04, 1.45],
    ['161125', '易方达标普500指数A', '指数型-海外股票', '标普500', 0.80, 0.20, '0.12%', 21.36, 6.54, 5.97, 3.60, 0.74],
    # 主动管理型QDII基金 (重仓苹果/科技)
    ['012920', '易方达全球成长精选混合A', '混合型-QDII', '全球科技成长', 1.20, 0.20, '0.15%', 202.49, 77.89, 46.00, 20.19, 2.32],
    ['016664', '天弘全球高端制造混合A', '混合型-QDII', '全球高端制造', 1.20, 0.20, '0.15%', 187.55, 97.24, 44.34, 22.07, 2.20],
    ['017730', '嘉实全球产业升级股票A', '股票型-QDII', '全球产业升级', 1.20, 0.20, '0.15%', 157.98, 78.03, 46.02, 22.06, 1.64],
    ['006373', '国富全球科技互联混合A', '混合型-QDII', '全球科技互联', 1.20, 0.20, '0.15%', 111.16, 82.32, 41.52, 23.38, 0.76],
    ['005698', '华夏全球科技先锋混合A', '混合型-QDII', '全球科技先锋', 1.20, 0.20, '0.15%', 104.29, 71.23, 45.60, 15.81, 1.07],
    ['100055', '富国全球科技互联网股票A', '股票型-QDII', '全球科技互联网', 1.20, 0.20, '0.15%', 94.86, 49.52, 39.82, 16.10, 1.44],
    ['270023', '广发全球精选股票A', '股票型-QDII', '全球精选', 1.20, 0.20, '0.16%', 68.47, 47.19, 29.40, 12.79, 1.24],
    ['001668', '汇添富全球移动互联混合A', '混合型-QDII', '全球移动互联', 1.20, 0.20, '0.16%', 40.44, 23.14, 22.04, 9.18, -0.09],
    ['017436', '华宝纳斯达克精选股票A', '股票型-QDII', '纳斯达克精选', 1.20, 0.20, '0.15%', 32.17, 7.48, 13.67, 3.16, -0.09],
]

# 颜色定义
light_blue = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
light_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
light_yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# 数据行
for row_idx, fund in enumerate(funds_data, 3):
    # 交替行颜色
    if row_idx <= 9:  # 被动指数型
        fill = light_blue if row_idx % 2 == 1 else white_fill
    else:  # 主动管理型
        fill = light_green if row_idx % 2 == 1 else white_fill

    for col_idx, value in enumerate(fund, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='微软雅黑', size=10)

        # 收益率列添加条件格式
        if col_idx >= 8 and isinstance(value, (int, float)):
            if value > 50:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='FF0000')
            elif value > 20:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='FF6600')
            elif value > 0:
                cell.font = Font(name='微软雅黑', size=10, color='008000')
            else:
                cell.font = Font(name='微软雅黑', size=10, color='FF0000')

# 设置列宽
col_widths = [12, 30, 18, 18, 15, 15, 18, 12, 12, 12, 12, 12]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 添加分隔行
ws.merge_cells('A10:L10')
ws['A10'] = '--- 被动指数型基金 (跟踪纳斯达克100/标普500) ---'
ws['A10'].font = Font(name='微软雅黑', size=10, bold=True, color='2E75B6')
ws['A10'].alignment = Alignment(horizontal='center')

ws.merge_cells('A18:L18')
ws['A18'] = '--- 主动管理型QDII基金 (重仓苹果/全球科技) ---'
ws['A18'].font = Font(name='微软雅黑', size=10, bold=True, color='548235')
ws['A18'].alignment = Alignment(horizontal='center')

# ============ Sheet 2: 费率详细对比 ============
ws2 = wb.create_sheet("费率详细对比")

ws2.merge_cells('A1:G1')
ws2['A1'] = '苹果概念基金费率详细对比'
ws2['A1'].font = title_font
ws2['A1'].fill = title_fill
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 40

fee_headers = ['基金代码', '基金名称', '管理费率(%/年)', '托管费率(%/年)', '原申购费率', '支付宝优惠费率', '赎回费率(7天内)']
for col, header in enumerate(fee_headers, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

fee_data = [
    ['270042', '广发纳斯达克100ETF联接A', 0.80, 0.20, '1.30%', '0.13%', '1.50%'],
    ['006479', '广发纳斯达克100ETF联接C', 0.80, 0.20, '0.00%', '0.00%', '1.50%'],
    ['040046', '华安纳斯达克100ETF联接A', 0.60, 0.20, '1.20%', '0.12%', '1.50%'],
    ['015299', '华夏纳斯达克100ETF联接A', 0.60, 0.20, '1.20%', '0.12%', '1.50%'],
    ['000834', '大成纳斯达克100ETF联接A', 0.80, 0.20, '1.20%', '0.12%', '1.50%'],
    ['019547', '招商纳斯达克100ETF联接A', 0.50, 0.15, '1.20%', '0.12%', '1.50%'],
    ['161125', '易方达标普500指数A', 0.80, 0.20, '1.20%', '0.12%', '1.50%'],
    ['012920', '易方达全球成长精选混合A', 1.20, 0.20, '1.50%', '0.15%', '1.50%'],
    ['016664', '天弘全球高端制造混合A', 1.20, 0.20, '1.50%', '0.15%', '1.50%'],
    ['017730', '嘉实全球产业升级股票A', 1.20, 0.20, '1.50%', '0.15%', '1.50%'],
    ['006373', '国富全球科技互联混合A', 1.20, 0.20, '1.50%', '0.15%', '1.50%'],
    ['005698', '华夏全球科技先锋混合A', 1.20, 0.20, '1.50%', '0.15%', '1.50%'],
    ['100055', '富国全球科技互联网股票A', 1.20, 0.20, '1.50%', '0.15%', '1.50%'],
    ['270023', '广发全球精选股票A', 1.20, 0.20, '1.60%', '0.16%', '1.50%'],
    ['001668', '汇添富全球移动互联混合A', 1.20, 0.20, '1.60%', '0.16%', '1.50%'],
    ['017436', '华宝纳斯达克精选股票A', 1.20, 0.20, '1.50%', '0.15%', '1.50%'],
]

for row_idx, data in enumerate(fee_data, 3):
    fill = light_blue if row_idx % 2 == 1 else white_fill
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='微软雅黑', size=10)

for i, width in enumerate([12, 30, 15, 15, 15, 18, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ============ Sheet 3: 投资建议 ============
ws3 = wb.create_sheet("投资建议与买入时机")

ws3.merge_cells('A1:F1')
ws3['A1'] = '苹果发布会前基金定投策略建议'
ws3['A1'].font = title_font
ws3['A1'].fill = title_fill
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

# 苹果发布会时间线
ws3.merge_cells('A3:F3')
ws3['A3'] = '一、苹果2026年发布会时间线 (预计)'
ws3['A3'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')

events = [
    ['时间', '事件', '影响', '建议操作'],
    ['2026年6月8-12日(预计)', 'WWDC 2026 全球开发者大会', '发布iOS 20/macOS 17等新系统,可能发布新硬件', '发布会前1-2周为最佳买入窗口'],
    ['2026年9月(预计)', '秋季发布会', '发布iPhone 18系列', '7-8月为预期炒作期'],
    ['2026年10月(预计)', 'Mac/iPad发布会', '发布新Mac和iPad', '9月为买入窗口'],
]

for row_idx, event in enumerate(events, 4):
    for col_idx, value in enumerate(event, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 4:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = light_yellow if row_idx % 2 == 0 else white_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 买入窗口分析
ws3.merge_cells('A9:F9')
ws3['A9'] = '二、最佳买入窗口分析'
ws3['A9'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')

strategy = [
    ['策略', '时间窗口', '预期收益', '风险等级', '适合人群', '推荐基金类型'],
    ['提前布局', '发布会前30-45天', '中等(5-15%)', '中等', '有经验的投资者', '主动管理型QDII'],
    ['趋势跟踪', '发布会前15-30天', '较高(10-25%)', '较高', '风险承受能力较强', '纳斯达克100指数'],
    ['短线博弈', '发布会前7-15天', '高(15-30%)', '高', '激进型投资者', '科技主题QDII'],
    ['稳健定投', '每月固定日期', '稳定(年化15-30%)', '低', '所有投资者', '标普500/纳斯达克100'],
]

for row_idx, data in enumerate(strategy, 10):
    for col_idx, value in enumerate(data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 10:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = light_green if row_idx % 2 == 0 else white_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 推荐组合
ws3.merge_cells('A16:F16')
ws3['A16'] = '三、推荐基金组合 (按风险偏好)'
ws3['A16'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')

portfolio = [
    ['风险偏好', '推荐组合', '基金代码', '配置比例', '理由', '预期年化收益'],
    ['保守型', '纳斯达克100+标普500', '270042+161125', '70%+30%', '分散风险,跟踪大盘', '15-25%'],
    ['稳健型', '纳斯达克100+全球科技', '040046+006373', '60%+40%', '平衡收益与风险', '20-35%'],
    ['积极型', '全球科技+高端制造', '012920+016664', '50%+50%', '追求高收益', '30-50%'],
    ['激进型', '全球成长精选', '012920', '100%', '集中持仓,高收益高风险', '40-80%'],
]

for row_idx, data in enumerate(portfolio, 17):
    for col_idx, value in enumerate(data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 17:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = light_blue if row_idx % 2 == 1 else white_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 定投建议
ws3.merge_cells('A23:F23')
ws3['A23'] = '四、定投操作建议'
ws3['A23'].font = Font(name='微软雅黑', size=12, bold=True, color='1F4E79')

tips = [
    '1. 定投频率: 建议每周或每两周定投一次,避免一次性大额买入',
    '2. 买入时机: WWDC 2026预计6月8-12日,建议从5月中旬开始分批建仓',
    '3. 止盈策略: 发布会后1-2周内,若收益达15-20%可考虑部分止盈',
    '4. 持有周期: 建议至少持有3-6个月,享受苹果新品发布后的业绩增长',
    '5. 风险控制: 单只基金仓位不超过总资产的30%,注意分散投资',
    '6. 汇率对冲: QDII基金受人民币汇率影响,可适当配置人民币份额',
    '7. 费率优化: C类份额适合短期持有(无申购费),A类份额适合长期持有',
]

for row_idx, tip in enumerate(tips, 24):
    ws3.merge_cells(f'A{row_idx}:F{row_idx}')
    cell = ws3.cell(row=row_idx, column=1, value=tip)
    cell.font = Font(name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 设置列宽
for i, width in enumerate([18, 25, 20, 15, 30, 18], 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

# 保存文件
output_path = r'c:\Users\gdx\soulquad-web\src\苹果概念基金分析_20260526.xlsx'
wb.save(output_path)
print(f"Excel文件已生成: {output_path}")
